from __future__ import annotations

import csv
from io import BytesIO, StringIO
from pathlib import PurePosixPath
from typing import Any

import markdown as markdown_renderer
from openpyxl import load_workbook

from app.services.file_reference_access import (
    AuthorizedFileSpaceContext,
    read_authorized_bytes_reference,
    read_authorized_text_reference,
    resolve_authorized_file_reference,
)
from app.services.file_spaces.service import FileSpaceService, FileSpaceValidationError

RAW_TEXT_REPRESENTATION = "raw_text"
MARKDOWN_RENDERED_REPRESENTATION = "markdown_rendered"
DELIMITED_TABLE_REPRESENTATION = "delimited_table"
WORKBOOK_REPRESENTATION = "workbook"
BINARY_META_REPRESENTATION = "binary_meta"

_MARKDOWN_EXTENSIONS = {"md", "markdown", "mdown"}
_MARKDOWN_MIME_TYPES = {"text/markdown", "text/x-markdown", "application/markdown"}
_DELIMITED_EXTENSIONS = {"csv", "tsv"}
_DELIMITED_MIME_TYPES = {
    "text/csv",
    "application/csv",
    "text/tab-separated-values",
    "application/tab-separated-values",
    "text/tsv",
}
_XLSX_EXTENSIONS = {"xlsx"}
_XLSX_MIME_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel.sheet.macroenabled.12",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.template",
}


def _file_extension(path: str) -> str:
    return PurePosixPath(str(path or "")).suffix.lower().lstrip(".")


def _mime_type(entry: Any, revision: Any | None) -> str:
    raw_value = getattr(revision, "mime_type", None) or getattr(entry, "mime_type", None) or ""
    return str(raw_value).split(";", 1)[0].strip().lower()


def _is_markdown(entry: Any, revision: Any | None) -> bool:
    return _file_extension(getattr(entry, "path", "")) in _MARKDOWN_EXTENSIONS or _mime_type(entry, revision) in _MARKDOWN_MIME_TYPES


def _is_delimited(entry: Any, revision: Any | None) -> bool:
    return _file_extension(getattr(entry, "path", "")) in _DELIMITED_EXTENSIONS or _mime_type(entry, revision) in _DELIMITED_MIME_TYPES


def _is_xlsx(entry: Any, revision: Any | None) -> bool:
    return _file_extension(getattr(entry, "path", "")) in _XLSX_EXTENSIONS or _mime_type(entry, revision) in _XLSX_MIME_TYPES


def _category(entry: Any, revision: Any | None) -> str:
    if _is_delimited(entry, revision):
        return "delimited"
    if _is_markdown(entry, revision):
        return "markdown"
    if _is_xlsx(entry, revision):
        return "workbook"
    if bool(getattr(entry, "is_text", False)):
        return "text"
    return "binary"


def list_supported_representations(entry: Any, revision: Any | None) -> list[dict[str, Any]]:
    representations: list[dict[str, Any]] = []
    if bool(getattr(entry, "is_text", False)):
        representations.append({"id": RAW_TEXT_REPRESENTATION, "label": "Raw text"})
    if _is_markdown(entry, revision):
        representations.append({"id": MARKDOWN_RENDERED_REPRESENTATION, "label": "Rendered Markdown"})
    if _is_delimited(entry, revision):
        representations.append({"id": DELIMITED_TABLE_REPRESENTATION, "label": "Delimited table"})
    if _is_xlsx(entry, revision):
        representations.append({"id": WORKBOOK_REPRESENTATION, "label": "Workbook"})
    representations.append({"id": BINARY_META_REPRESENTATION, "label": "Binary metadata"})
    return representations


def _slice_text_content(
    *,
    path: str,
    content: str,
    start_line_raw: Any,
    end_line_raw: Any,
    include_line_numbers: bool,
) -> dict[str, Any]:
    lines = content.splitlines()
    total_lines = len(lines)
    if total_lines == 0:
        start_line = 1
        end_line = 0
    elif start_line_raw is not None and end_line_raw is not None:
        start_line = int(start_line_raw)
        end_line = int(end_line_raw)
    elif start_line_raw is not None:
        start_line = int(start_line_raw)
        end_line = min(total_lines, start_line + 199)
    elif end_line_raw is not None:
        end_line = int(end_line_raw)
        start_line = max(1, end_line - 199)
    elif total_lines <= 300:
        start_line = 1
        end_line = total_lines
    else:
        start_line = 1
        end_line = min(total_lines, 200)
    if total_lines > 0 and (start_line < 1 or end_line < start_line or end_line > total_lines):
        raise FileSpaceValidationError("start_line and end_line must define a valid inclusive range")
    selected_lines = [] if total_lines == 0 else lines[start_line - 1 : end_line]
    payload: dict[str, Any] = {
        "path": path,
        "content": "\n".join(selected_lines),
        "total_lines": total_lines,
        "start_line": start_line,
        "end_line": end_line,
        "truncated": total_lines > 0 and (start_line > 1 or end_line < total_lines),
    }
    if include_line_numbers:
        payload["numbered_content"] = "\n".join(f"{line_no}: {line}" for line_no, line in enumerate(selected_lines, start=start_line))
    return payload


def _coerce_max_rows(value: Any, *, default: int = 200, maximum: int = 1000) -> int:
    if value in (None, ""):
        return default
    try:
        parsed = int(value)
    except Exception as exc:
        raise FileSpaceValidationError("max_rows must be an integer") from exc
    return max(1, min(parsed, maximum))


def _detect_delimiter(content: str, *, path: str, mime_type: str) -> str:
    if _file_extension(path) == "tsv" or "tab-separated-values" in mime_type:
        return "\t"
    try:
        dialect = csv.Sniffer().sniff(content[:4096] or ",", delimiters=[",", ";", "\t"])
        return str(dialect.delimiter or ",")
    except Exception:
        return ","


def _normalize_tabular_rows(rows: list[list[Any]]) -> tuple[list[list[str]], int]:
    normalized_rows = [[("" if cell is None else str(cell)) for cell in row] for row in rows]
    column_count = max((len(row) for row in normalized_rows), default=0)
    return normalized_rows, column_count


async def inspect_file(
    ctx: AuthorizedFileSpaceContext,
    *,
    path: str,
) -> dict[str, Any]:
    file_ref = await resolve_authorized_file_reference(ctx, path=path)
    return {
        "entry": FileSpaceService.serialize_entry(file_ref.entry),
        "revision": FileSpaceService.serialize_revision(file_ref.revision) if file_ref.revision is not None else None,
        "file_kind": {
            "category": _category(file_ref.entry, file_ref.revision),
            "mime_type": _mime_type(file_ref.entry, file_ref.revision) or None,
            "extension": _file_extension(file_ref.path) or None,
            "is_text": bool(getattr(file_ref.entry, "is_text", False)),
        },
        "representations": list_supported_representations(file_ref.entry, file_ref.revision),
    }


async def read_representation(
    ctx: AuthorizedFileSpaceContext,
    *,
    path: str,
    representation: str,
    options: dict[str, Any] | None = None,
) -> dict[str, Any]:
    normalized_representation = str(representation or "").strip()
    normalized_options = options if isinstance(options, dict) else {}

    if normalized_representation == RAW_TEXT_REPRESENTATION:
        file_ref, content = await read_authorized_text_reference(ctx, path=path)
        return {
            "representation": RAW_TEXT_REPRESENTATION,
            "entry": FileSpaceService.serialize_entry(file_ref.entry),
            "revision": FileSpaceService.serialize_revision(file_ref.revision),
            **_slice_text_content(
                path=file_ref.path,
                content=content,
                start_line_raw=normalized_options.get("start_line") or normalized_options.get("startLine"),
                end_line_raw=normalized_options.get("end_line") or normalized_options.get("endLine"),
                include_line_numbers=bool(
                    normalized_options.get("include_line_numbers") or normalized_options.get("includeLineNumbers")
                ),
            ),
        }

    if normalized_representation == MARKDOWN_RENDERED_REPRESENTATION:
        file_ref, content = await read_authorized_text_reference(ctx, path=path)
        if not _is_markdown(file_ref.entry, file_ref.revision):
            raise FileSpaceValidationError("markdown_rendered is not supported for this file")
        return {
            "representation": MARKDOWN_RENDERED_REPRESENTATION,
            "entry": FileSpaceService.serialize_entry(file_ref.entry),
            "revision": FileSpaceService.serialize_revision(file_ref.revision),
            "html": markdown_renderer.markdown(content, extensions=["extra", "tables", "fenced_code"]),
        }

    if normalized_representation == DELIMITED_TABLE_REPRESENTATION:
        file_ref, content = await read_authorized_text_reference(ctx, path=path)
        if not _is_delimited(file_ref.entry, file_ref.revision):
            raise FileSpaceValidationError("delimited_table is not supported for this file")
        delimiter = _detect_delimiter(content, path=file_ref.path, mime_type=_mime_type(file_ref.entry, file_ref.revision))
        rows = list(csv.reader(StringIO(content), delimiter=delimiter))
        max_rows = _coerce_max_rows(normalized_options.get("max_rows") or normalized_options.get("maxRows"))
        normalized_rows, column_count = _normalize_tabular_rows(rows[:max_rows])
        return {
            "representation": DELIMITED_TABLE_REPRESENTATION,
            "entry": FileSpaceService.serialize_entry(file_ref.entry),
            "revision": FileSpaceService.serialize_revision(file_ref.revision),
            "delimiter": delimiter,
            "rows": normalized_rows,
            "row_count": len(rows),
            "displayed_row_count": len(normalized_rows),
            "column_count": column_count,
            "truncated": len(rows) > max_rows,
        }

    if normalized_representation == WORKBOOK_REPRESENTATION:
        file_ref, payload = await read_authorized_bytes_reference(ctx, path=path)
        if not _is_xlsx(file_ref.entry, file_ref.revision):
            raise FileSpaceValidationError("workbook is not supported for this file")
        workbook = load_workbook(filename=BytesIO(payload), read_only=True, data_only=True)
        sheet_names = list(workbook.sheetnames)
        selected_sheet = str(normalized_options.get("sheet") or "").strip() or (sheet_names[0] if sheet_names else "")
        if not selected_sheet or selected_sheet not in workbook.sheetnames:
            raise FileSpaceValidationError("requested sheet is not available")
        worksheet = workbook[selected_sheet]
        max_rows = _coerce_max_rows(normalized_options.get("max_rows") or normalized_options.get("maxRows"))
        rows: list[list[Any]] = []
        truncated = False
        for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            if row_index > max_rows:
                truncated = True
                break
            rows.append(list(row))
        normalized_rows, column_count = _normalize_tabular_rows(rows)
        return {
            "representation": WORKBOOK_REPRESENTATION,
            "entry": FileSpaceService.serialize_entry(file_ref.entry),
            "revision": FileSpaceService.serialize_revision(file_ref.revision),
            "sheet_names": sheet_names,
            "sheet": selected_sheet,
            "rows": normalized_rows,
            "displayed_row_count": len(normalized_rows),
            "column_count": column_count,
            "truncated": truncated,
        }

    if normalized_representation == BINARY_META_REPRESENTATION:
        file_ref = await resolve_authorized_file_reference(ctx, path=path)
        return {
            "representation": BINARY_META_REPRESENTATION,
            "entry": FileSpaceService.serialize_entry(file_ref.entry),
            "revision": FileSpaceService.serialize_revision(file_ref.revision) if file_ref.revision is not None else None,
        }

    raise FileSpaceValidationError("unsupported representation")
